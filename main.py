import os
import io
import re
import json
import uuid
import time
import asyncio
import datetime
from collections import OrderedDict

import requests
import numpy as np
import pandas as pd
from rapidfuzz import process, fuzz
from fastapi import FastAPI, UploadFile, File, HTTPException, Body
from fastapi.middleware.cors import CORSMiddleware
from fastapi.staticfiles import StaticFiles
from fastapi.responses import StreamingResponse

app = FastAPI(title="DotDash")

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_methods=["*"],
    allow_headers=["*"],
)

# In-memory store: rb_session_id -> {"df": DataFrame, "sheet_name": str, "weeks": [...],
#                                     "query_cache": {...}}
# "shared" is a fixed key: a team-wide dashboard that everyone sees the same data for,
# refreshed from a Google Sheet link rather than a personal file upload.
RB_SESSIONS = {}
SHARED_SESSION_ID = "shared"
SHARED_AUTO_REFRESH_SECONDS = 30

# Subscribers for the shared dashboard's Server-Sent Events stream.
SHARED_SUBSCRIBERS: list = []

DEFAULT_MODELS = {
    "anthropic": "claude-sonnet-4-5",
    "openai": "gpt-4o-mini",
    "gemini": "gemini-2.5-flash",
}

# ---------------------------------------------------------------------------
# Small bounded cache for NL -> intent parsing, so repeated/near-identical
# questions ("show me farming team", asked again a minute later) don't
# re-hit the LLM. Keyed on session + exact prompt text + provider/model.
# Cleared implicitly whenever a session is replaced (new upload/refresh),
# since the session_id changes for uploads and NL intents don't depend on
# the underlying numbers — only on category names, which change rarely.
# ---------------------------------------------------------------------------
NL_INTENT_CACHE: "OrderedDict" = OrderedDict()
NL_CACHE_MAX = 500


def nl_cache_get(key):
    if key in NL_INTENT_CACHE:
        NL_INTENT_CACHE.move_to_end(key)
        return NL_INTENT_CACHE[key]
    return None


def nl_cache_set(key, value):
    NL_INTENT_CACHE[key] = value
    NL_INTENT_CACHE.move_to_end(key)
    if len(NL_INTENT_CACHE) > NL_CACHE_MAX:
        NL_INTENT_CACHE.popitem(last=False)


def call_llm(provider: str, api_key: str, model: str, system_prompt: str, user_msg: str) -> str:
    """Calls whichever provider the user picked, with their own key. Returns raw text."""
    provider = (provider or "").lower()
    model = model or DEFAULT_MODELS.get(provider)

    if provider == "anthropic":
        import anthropic
        client = anthropic.Anthropic(api_key=api_key)
        resp = client.messages.create(
            model=model,
            max_tokens=1000,
            system=system_prompt,
            messages=[{"role": "user", "content": user_msg}],
        )
        return "".join(b.text for b in resp.content if b.type == "text")

    elif provider == "openai":
        from openai import OpenAI
        client = OpenAI(api_key=api_key)
        resp = client.chat.completions.create(
            model=model,
            messages=[
                {"role": "system", "content": system_prompt},
                {"role": "user", "content": user_msg},
            ],
        )
        return resp.choices[0].message.content

    elif provider == "gemini":
        import google.generativeai as genai
        genai.configure(api_key=api_key)
        gmodel = genai.GenerativeModel(model, system_instruction=system_prompt)
        resp = gmodel.generate_content(user_msg)
        return resp.text

    else:
        raise ValueError(f"Unknown provider: {provider}")


@app.get("/api/health")
def health():
    return {"status": "ok"}


# ---------------------------------------------------------------------------
# Auto-detects a weekly merchant revenue tracking sheet (columns like
# W0, W1, W_2... + Regional Head) inside a multi-sheet workbook, and
# computes week-over-week revenue bridge KPIs.
# ---------------------------------------------------------------------------

RB_FIELD_PATTERNS = {
    "merchant": ["merchant name"],
    "store": ["store #", "store#"],
    "bucket": ["bucket"],
    "city": ["head office city", "city"],
    "region": ["region"],
    "kam": ["actual kam", "kam"],
    "rh": ["regional head"],
    "team": ["team"],
    "segment": ["type of merchant"],
    "product": ["product"],
    "received": ["received"],
    "spillover": ["spillover amount"],
}


def find_weekly_sheet(wb):
    """Scans all sheets for ones with week columns (W0, W1, W_2, ...) and a
    Regional Head column, then scores candidates to pick the true master
    sheet (most distinct teams, least duplicate merchant rows) rather than
    just the first match — workbooks often have several per-team or
    per-purpose sheets that also happen to match the pattern."""
    candidates = []
    for sn in wb.sheetnames:
        ws = wb[sn]
        max_scan_row = min(6, ws.max_row + 1)
        for r in range(1, max_scan_row):
            row_vals = [str(c.value).strip() if c.value else "" for c in ws[r]]
            has_week = any(re.match(r'^W_?\d+\b', v, re.I) for v in row_vals)
            has_rh = any("regional head" in v.lower() for v in row_vals)
            if has_week and has_rh:
                candidates.append((sn, r))
                break

    if not candidates:
        return None, None
    if len(candidates) == 1:
        return candidates[0]

    scored = []
    for sn, header_row in candidates:
        ws = wb[sn]
        headers = [(c, str(ws.cell(row=header_row, column=c).value or "").strip()) for c in range(1, ws.max_column + 1)]
        team_col = next((c for c, h in headers if h.lower() == "team"), None)
        merchant_col = next((c for c, h in headers if h.lower() == "merchant name"), None)
        if not merchant_col:
            continue

        teams, names, total = set(), [], 0
        for r in range(header_row + 1, ws.max_row + 1):
            mv = ws.cell(row=r, column=merchant_col).value
            if mv:
                total += 1
                names.append(str(mv).strip())
                if team_col:
                    tv = ws.cell(row=r, column=team_col).value
                    if tv:
                        teams.add(str(tv).strip().upper())

        uniqueness = len(set(names)) / total if total else 0
        scored.append((len(teams) >= 4, uniqueness, len(teams), total, sn, header_row))

    # Prefer sheets with a real spread of teams (>=4), then cleanest data
    # (least duplicate merchant rows), then more distinct teams, then more rows.
    scored.sort(key=lambda x: (x[0], x[1], x[2], x[3]), reverse=True)
    *_, sn, header_row = scored[0]
    return sn, header_row


def detect_columns(headers):
    """headers: list of (index, header_string). Returns dict of field -> column index,
    and week_cols dict of week_number -> column index (rightmost occurrence wins)."""
    field_cols = {}
    week_cols = {}
    for idx, h in headers:
        h_clean = h.strip()
        h_lower = h_clean.lower()

        m = re.match(r'^W_?(\d+)\b', h_clean, re.I)
        if m:
            week_cols[int(m.group(1))] = idx
            continue

        for field, patterns in RB_FIELD_PATTERNS.items():
            for p in patterns:
                if p == h_lower or (p in h_lower and field != "region"):
                    field_cols[field] = idx
                    break
                if field == "region" and h_lower == "region":
                    field_cols[field] = idx
                    break

    return field_cols, week_cols


def _normalize_bridge_df(df: pd.DataFrame, week_cols) -> pd.DataFrame:
    """Shared cleanup used by both the xlsx and CSV loaders: coerce numeric
    columns, fill/trim text columns, and collapse casing variants of the
    same category (e.g. 'Hunting' / 'HUNTING') to one canonical label."""
    for wk in week_cols:
        col = f"W{wk}"
        if col in df.columns:
            df[col] = pd.to_numeric(df[col], errors="coerce").fillna(0)
    if "store" in df.columns:
        df["store"] = pd.to_numeric(df["store"], errors="coerce").fillna(0)
    for txt_field in ["team", "segment", "bucket", "city", "region", "kam", "rh", "merchant", "product"]:
        if txt_field in df.columns:
            df[txt_field] = df[txt_field].fillna("Unknown").astype(str).str.strip()
            df.loc[df[txt_field] == "", txt_field] = "Unknown"
    for txt_field in ["team", "segment", "city", "region", "product"]:
        if txt_field in df.columns:
            canonical = {}
            for val in df[txt_field]:
                key = val.upper()
                if key not in canonical:
                    canonical[key] = val
            df[txt_field] = df[txt_field].map(lambda v: canonical[v.upper()])
    return df


def load_revenue_bridge_workbook(wb, sheet_name: str, header_row: int):
    """Takes an already-opened workbook (avoids re-parsing the file a second
    time) and extracts the weekly revenue bridge sheet into a DataFrame."""
    ws = wb[sheet_name]
    headers = [(c + 1, str(ws.cell(row=header_row, column=c + 1).value or "")) for c in range(ws.max_column)]
    headers = [(i, h) for i, h in headers if h.strip()]
    field_cols, week_cols = detect_columns(headers)

    required = ["merchant", "rh"]
    missing = [f for f in required if f not in field_cols]
    if missing or not week_cols:
        raise ValueError(
            f"Found sheet '{sheet_name}' but couldn't identify required columns: "
            f"{missing or []} {'(no week columns found)' if not week_cols else ''}"
        )

    # Map 0-indexed column position -> ("field", name) or ("week", week_num),
    # then bulk-read rows with iter_rows(values_only=True) instead of making
    # one .cell() lookup per field per row — much faster on 15k-25k+ rows.
    pos_to_field = {idx - 1: ("field", field) for field, idx in field_cols.items()}
    pos_to_field.update({idx - 1: ("week", wk) for wk, idx in week_cols.items()})
    merchant_field_present = "merchant" in field_cols

    data_rows = []
    for row_tuple in ws.iter_rows(min_row=header_row + 1, values_only=True):
        row = {}
        empty = True
        for pos, val in enumerate(row_tuple):
            mapping = pos_to_field.get(pos)
            if mapping is None:
                continue
            kind, key = mapping
            if kind == "field":
                row[key] = val
            else:
                row[f"W{key}"] = val if isinstance(val, (int, float)) else 0
            if val not in (None, "", 0):
                empty = False
        if not empty and merchant_field_present and row.get("merchant"):
            data_rows.append(row)

    df = pd.DataFrame(data_rows)
    df = _normalize_bridge_df(df, week_cols)

    return {
        "df": df,
        "sheet_name": sheet_name,
        "weeks": sorted(week_cols.keys()),
        "has_received": "received" in df.columns,
        "has_spillover": "spillover" in df.columns,
    }


# ---------------------------------------------------------------------------
# Generic fallback: any sheet that doesn't match the weekly revenue bridge
# pattern still gets a real dashboard — an auto-generated overview (one
# chart/stat per column) plus natural-language Q&A over the raw data.
# ---------------------------------------------------------------------------

def load_generic_workbook(wb):
    """Picks the sheet with the most data, guesses its header row, and loads it
    into a DataFrame without assuming any particular schema."""
    best_sheet, best_rows = wb.sheetnames[0], -1
    for sn in wb.sheetnames:
        ws = wb[sn]
        if ws.max_row > best_rows:
            best_sheet, best_rows = sn, ws.max_row
    ws = wb[best_sheet]

    header_row, best_score = 1, -1
    for r in range(1, min(6, ws.max_row + 1)):
        vals = [ws.cell(row=r, column=c).value for c in range(1, min(ws.max_column, 60) + 1)]
        score = sum(1 for v in vals if isinstance(v, str) and v.strip())
        if score > best_score:
            best_score, header_row = score, r

    seen, headers = {}, []
    for c in range(1, ws.max_column + 1):
        h = ws.cell(row=header_row, column=c).value
        h = str(h).strip() if h else f"column_{c}"
        if h in seen:
            seen[h] += 1
            h = f"{h}_{seen[h]}"
        else:
            seen[h] = 0
        headers.append(h)

    rows = []
    for row_tuple in ws.iter_rows(min_row=header_row + 1, values_only=True):
        row, empty = {}, True
        for c, h in enumerate(headers):
            v = row_tuple[c] if c < len(row_tuple) else None
            row[h] = v
            if v not in (None, ""):
                empty = False
        if not empty:
            rows.append(row)

    if not rows:
        raise ValueError(f"Sheet '{best_sheet}' doesn't have any data rows yet — nothing to build a dashboard from.")

    df = pd.DataFrame(rows)
    for col in df.columns:
        sample = df[col].dropna()
        if not sample.empty and sample.apply(lambda v: isinstance(v, (datetime.date, datetime.datetime))).all():
            # Dates/datetimes: keep them human-readable, never treat as numeric
            df[col] = df[col].apply(
                lambda v: v.strftime("%Y-%m-%d") if isinstance(v, (datetime.date, datetime.datetime)) else v
            )
            continue
        converted = pd.to_numeric(df[col], errors="coerce")
        if converted.notna().sum() >= 0.7 * len(df) and len(df) > 0:
            df[col] = converted

    return {"df": df, "sheet_name": best_sheet}


def build_generic_overview(df: pd.DataFrame, max_widgets: int = 8):
    widgets = []
    for col in list(df.columns)[:max_widgets]:
        s = df[col].dropna()
        if s.empty:
            continue
        if pd.api.types.is_numeric_dtype(s):
            bins = min(8, max(3, s.nunique()))
            try:
                counts, edges = np.histogram(s, bins=bins)
            except Exception:
                continue
            data = [{"bucket": f"{edges[i]:.1f}–{edges[i+1]:.1f}", "count": int(counts[i])} for i in range(len(counts))]
            widgets.append({
                "title": f"{col} — distribution",
                "chart_type": "bar", "chart_x": "bucket", "chart_y": "count",
                "data": data,
                "stat": {
                    "mean": round(float(s.mean()), 2), "min": round(float(s.min()), 2),
                    "max": round(float(s.max()), 2), "sum": round(float(s.sum()), 2),
                },
            })
        else:
            vc = s.astype(str).value_counts().head(8)
            data = [{"label": str(idx), "value": int(v)} for idx, v in vc.items()]
            widgets.append({
                "title": f"{col} — top values",
                "chart_type": "bar", "chart_x": "label", "chart_y": "value",
                "data": data,
            })
    return widgets


def load_workbook_smart(content: bytes) -> dict:
    """Tries the revenue-bridge pattern first (W0/W1... + Regional Head columns);
    falls back to a generic auto-dashboard for anything else. Only opens the
    workbook once and reuses it for whichever path is taken."""
    import openpyxl
    wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
    sheet_name, header_row = find_weekly_sheet(wb)

    if sheet_name is not None:
        try:
            parsed = load_revenue_bridge_workbook(wb, sheet_name, header_row)
            parsed["dashboard_type"] = "revenue_bridge"
            return parsed
        except ValueError:
            pass  # matched the pattern loosely but couldn't fully parse — fall through

    parsed = load_generic_workbook(wb)
    parsed["dashboard_type"] = "generic"
    parsed["overview_widgets"] = build_generic_overview(parsed["df"])
    parsed["columns"] = list(parsed["df"].columns)
    return parsed


def load_csv_smart(content: bytes) -> dict:
    """CSV equivalent of load_workbook_smart: tries to match the weekly
    revenue bridge column pattern by header name, falls back to generic."""
    raw_df = pd.read_csv(io.BytesIO(content))
    headers = [(i + 1, str(h)) for i, h in enumerate(raw_df.columns)]
    field_cols, week_cols = detect_columns(headers)

    if "merchant" in field_cols and "rh" in field_cols and week_cols:
        idx_to_colname = {i + 1: h for i, h in enumerate(raw_df.columns)}
        rename = {idx_to_colname[idx]: field for field, idx in field_cols.items()}
        rename.update({idx_to_colname[idx]: f"W{wk}" for wk, idx in week_cols.items()})
        df = raw_df.rename(columns=rename)
        keep_cols = [c for c in list(field_cols.keys()) + [f"W{wk}" for wk in week_cols] if c in df.columns]
        df = df[keep_cols].copy()
        df = _normalize_bridge_df(df, week_cols)
        return {
            "df": df,
            "sheet_name": "CSV upload",
            "weeks": sorted(week_cols.keys()),
            "has_received": "received" in df.columns,
            "has_spillover": "spillover" in df.columns,
            "dashboard_type": "revenue_bridge",
        }

    parsed = {"df": raw_df, "sheet_name": "CSV upload"}
    parsed["dashboard_type"] = "generic"
    parsed["overview_widgets"] = build_generic_overview(raw_df)
    parsed["columns"] = list(raw_df.columns)
    return parsed


def load_dataset_smart(content: bytes, filename: str = "") -> dict:
    if filename.lower().endswith(".csv"):
        return load_csv_smart(content)
    return load_workbook_smart(content)


def get_schema_summary(df: pd.DataFrame) -> str:
    lines = []
    for col in df.columns:
        dtype = str(df[col].dtype)
        sample = df[col].dropna().unique()[:5]
        sample_str = ", ".join(str(s) for s in sample)
        lines.append(f"- {col} ({dtype}): sample values -> {sample_str}")
    return "\n".join(lines)


def safe_exec_generic(code: str, df: pd.DataFrame):
    allowed_globals = {
        "pd": pd, "np": np, "df": df,
        "__builtins__": {
            "len": len, "sum": sum, "min": min, "max": max, "round": round,
            "sorted": sorted, "range": range, "list": list, "dict": dict,
            "str": str, "int": int, "float": float, "bool": bool,
            "enumerate": enumerate, "zip": zip, "abs": abs,
        },
    }
    local_vars = {}
    exec(code, allowed_globals, local_vars)
    return local_vars.get("result")


def generic_result_to_json(result):
    if isinstance(result, pd.DataFrame):
        r = result.reset_index(drop=True) if result.index.name is None else result.reset_index()
        return json.loads(r.to_json(orient="records", date_format="iso"))
    if isinstance(result, pd.Series):
        r = result.reset_index()
        r.columns = ["label", "value"]
        return json.loads(r.to_json(orient="records", date_format="iso"))
    if hasattr(result, "item"):
        try:
            return result.item()
        except Exception:
            pass
    return result


GENERIC_SYSTEM_PROMPT = """You are a data analyst assistant. You are given a pandas DataFrame called `df` and a question about it.

Respond with ONLY a JSON object (no markdown fences) with:
- "pandas_code": a short python snippet using `df` (and `pd`, `np`) that computes the answer and assigns it to a variable named `result`.
- "chart_type": one of "bar", "line", "pie", "scatter", "table", "none".
- "chart_x": column/index name for the x-axis, or null.
- "chart_y": column name for the y-axis, or null.
- "explanation": a short 1-3 sentence natural language answer, referencing actual values.

Rules: only pandas/numpy operations, no file I/O, no imports, no network calls. Keep pandas_code short and safe.
"""


def apply_filters(df: pd.DataFrame, filters: dict) -> pd.DataFrame:
    """filters: dict of field -> value, where value 'All'/'All Merchants'/None means no filter."""
    filtered = df.copy()
    no_filter_values = (None, "", "all", "all merchants")
    for field, value in filters.items():
        if value is None or str(value).strip().lower() in no_filter_values:
            continue
        if field in filtered.columns:
            filtered = filtered[filtered[field].astype(str).str.lower() == str(value).lower()]
    return filtered


def build_periods(weeks: list, weeks_per_month: int = 4, weeks_per_quarter: int = 13):
    """Groups sequential week numbers into Month/Quarter buckets. This is a defined
    convention (not derived from real calendar dates, since the sheet only has
    sequential week snapshots) — adjust weeks_per_month/quarter if your cadence differs."""
    weeks = sorted(weeks)
    months, quarters = {}, {}
    for w in weeks:
        months.setdefault(w // weeks_per_month + 1, []).append(w)
        quarters.setdefault(w // weeks_per_quarter + 1, []).append(w)

    month_list = [
        {"label": f"Month {m}", "prev_week": min(ws), "curr_week": max(ws)}
        for m, ws in sorted(months.items())
    ]
    quarter_list = [
        {"label": f"Q{q}", "prev_week": min(ws), "curr_week": max(ws)}
        for q, ws in sorted(quarters.items())
    ]
    return month_list, quarter_list


def compute_metric_bridge(df: pd.DataFrame, prev_week: int, curr_week: int, filters: dict, metric: str = "revenue"):
    """Generalized bridge: same Previous + New - Reduction = Current logic, applied to
    whichever metric is selected — revenue (₹ sum), merchants (distinct active count),
    or stores (store count)."""
    prev_col, curr_col = f"W{prev_week}", f"W{curr_week}"
    if prev_col not in df.columns or curr_col not in df.columns:
        raise ValueError(f"Week columns {prev_col}/{curr_col} not found.")
    if metric not in ("revenue", "merchants", "stores"):
        raise ValueError(f"Unknown metric: {metric}")

    filtered = apply_filters(df, filters)
    prev_active = filtered[prev_col] > 0
    curr_active = filtered[curr_col] > 0

    if metric == "revenue":
        delta = filtered[curr_col] - filtered[prev_col]
        new_val = float(delta.clip(lower=0).sum())
        reduction_val = float((-delta.clip(upper=0)).sum())
        previous_val = float(filtered[prev_col].sum())
        current_val = float(filtered[curr_col].sum())
    elif metric == "merchants":
        has_merchant = "merchant" in filtered.columns
        previous_val = float(filtered.loc[prev_active, "merchant"].nunique()) if has_merchant else float(prev_active.sum())
        current_val = float(filtered.loc[curr_active, "merchant"].nunique()) if has_merchant else float(curr_active.sum())
        new_mask = curr_active & ~prev_active
        churn_mask = prev_active & ~curr_active
        new_val = float(filtered.loc[new_mask, "merchant"].nunique()) if has_merchant else float(new_mask.sum())
        reduction_val = float(filtered.loc[churn_mask, "merchant"].nunique()) if has_merchant else float(churn_mask.sum())
    else:  # stores
        has_store = "store" in filtered.columns
        previous_val = float(filtered.loc[prev_active, "store"].sum()) if has_store else 0.0
        current_val = float(filtered.loc[curr_active, "store"].sum()) if has_store else 0.0
        new_mask = curr_active & ~prev_active
        churn_mask = prev_active & ~curr_active
        new_val = float(filtered.loc[new_mask, "store"].sum()) if has_store else 0.0
        reduction_val = float(filtered.loc[churn_mask, "store"].sum()) if has_store else 0.0

    subtotal = previous_val + new_val

    active = filtered[prev_active | curr_active]
    merchant_count = active["merchant"].nunique() if "merchant" in active.columns else len(active)
    store_count = float(active["store"].sum()) if "store" in active.columns else None
    received = float(filtered["received"].sum()) if "received" in filtered.columns else None
    spillover = float(filtered["spillover"].sum()) if "spillover" in filtered.columns else None

    rh_breakdown = []
    if "rh" in filtered.columns:
        if metric == "revenue":
            grp = filtered.groupby("rh").agg(
                value=(curr_col, "sum"),
                merchants=("merchant", "nunique") if "merchant" in filtered.columns else (curr_col, "count"),
                stores=("store", "sum") if "store" in filtered.columns else (curr_col, "count"),
            ).reset_index()
        elif metric == "merchants":
            grp = filtered[curr_active].groupby("rh").agg(
                value=("merchant", "nunique") if "merchant" in filtered.columns else (curr_col, "count"),
                merchants=("merchant", "nunique") if "merchant" in filtered.columns else (curr_col, "count"),
                stores=("store", "sum") if "store" in filtered.columns else (curr_col, "count"),
            ).reset_index()
        else:  # stores
            grp = filtered[curr_active].groupby("rh").agg(
                value=("store", "sum") if "store" in filtered.columns else (curr_col, "count"),
                merchants=("merchant", "nunique") if "merchant" in filtered.columns else (curr_col, "count"),
                stores=("store", "sum") if "store" in filtered.columns else (curr_col, "count"),
            ).reset_index()

        grp = grp[grp["value"] != 0].sort_values("value", ascending=False)
        max_val = grp["value"].abs().max() if len(grp) else 1
        for _, row in grp.iterrows():
            rh_breakdown.append({
                "rh": row["rh"],
                "current": round(float(row["value"]), 2),
                "merchants": int(row["merchants"]),
                "stores": int(row["stores"]) if pd.notna(row["stores"]) else 0,
                "bar_pct": round(abs(float(row["value"])) / max_val * 100, 1) if max_val else 0,
            })

    table_cols = [c for c in ["merchant", "store", "bucket", "city", "region", "product", "kam", "rh"] if c in filtered.columns]
    table_df = filtered[table_cols + [curr_col, prev_col]].copy() if table_cols else filtered[[curr_col, prev_col]].copy()
    if metric == "revenue":
        delta = filtered[curr_col] - filtered[prev_col]
        table_df["reduction"] = (-delta.clip(upper=0))
        table_df["new_opportunities"] = delta.clip(lower=0)
    else:
        table_df["reduction"] = churn_mask.astype(int)
        table_df["new_opportunities"] = new_mask.astype(int)
    table_df = table_df.rename(columns={curr_col: "current_week", prev_col: "previous_week"})
    table_df = table_df[(table_df["current_week"] != 0) | (table_df["previous_week"] != 0)]
    table_df = table_df.sort_values("current_week", ascending=False).head(500)
    table_rows = json.loads(table_df.to_json(orient="records"))

    return {
        "metric": metric,
        "previous_week_total": round(previous_val, 2),
        "new_opportunities": round(new_val, 2),
        "subtotal": round(subtotal, 2),
        "reduction": round(reduction_val, 2),
        "current_week_total": round(current_val, 2),
        "received": round(received, 2) if received is not None else None,
        "spillover": round(spillover, 2) if spillover is not None else None,
        "merchant_count": int(merchant_count),
        "store_count": int(store_count) if store_count is not None else None,
        "rh_breakdown": rh_breakdown,
        "table_rows": table_rows,
    }


def compute_trend(df: pd.DataFrame, weeks: list, filters: dict, metric: str = "revenue"):
    """Per-week totals (not a two-week bridge) across every week in the sheet,
    for the trend chart."""
    filtered = apply_filters(df, filters)
    points = []
    for wk in weeks:
        col = f"W{wk}"
        if col not in filtered.columns:
            continue
        active = filtered[col] > 0
        if metric == "revenue":
            val = float(filtered[col].sum())
        elif metric == "merchants":
            val = float(filtered.loc[active, "merchant"].nunique()) if "merchant" in filtered.columns else float(active.sum())
        else:  # stores
            val = float(filtered.loc[active, "store"].sum()) if "store" in filtered.columns else float(active.sum())
        points.append({"week": wk, "value": round(val, 2)})
    return points


def tab_counts(df: pd.DataFrame, field: str, week_col: str):
    if field not in df.columns:
        return []
    grp = df[df[week_col] != 0].groupby(field).agg(
        merchants=("merchant", "nunique") if "merchant" in df.columns else (week_col, "count"),
        stores=("store", "sum") if "store" in df.columns else (week_col, "count"),
    ).reset_index()
    return [
        {"name": row[field], "merchants": int(row["merchants"]), "stores": int(row["stores"]) if pd.notna(row["stores"]) else 0}
        for _, row in grp.iterrows()
    ]


def extract_sheet_id(url: str) -> str:
    m = re.search(r'/spreadsheets/d/([a-zA-Z0-9-_]+)', url)
    if not m:
        raise ValueError("That doesn't look like a Google Sheets link.")
    return m.group(1)


def fetch_gsheet_as_xlsx(url: str) -> bytes:
    """Downloads the whole workbook via Google's public export link. Requires the
    sheet to be shared as 'Anyone with the link can view'."""
    sheet_id = extract_sheet_id(url)
    export_url = f"https://docs.google.com/spreadsheets/d/{sheet_id}/export?format=xlsx"
    try:
        resp = requests.get(export_url, timeout=30, allow_redirects=True)
    except requests.RequestException as e:
        raise ValueError(f"Couldn't reach Google Sheets: {e}")

    if resp.status_code in (401, 403):
        raise ValueError(
            "Can't access this sheet. In Google Sheets, click Share > General access > "
            "'Anyone with the link' (Viewer), then try again."
        )
    if resp.status_code != 200 or not resp.content[:2] == b'PK':
        raise ValueError(
            "Couldn't download this as a spreadsheet. Double-check the link is a Google "
            "Sheets URL and it's shared as 'Anyone with the link can view'."
        )
    return resp.content


def _shared_summary(parsed: dict, rb_session_id: str) -> dict:
    df = parsed["df"]
    dashboard_type = parsed.get("dashboard_type", "revenue_bridge")
    base = {
        "rb_session_id": rb_session_id,
        "dashboard_type": dashboard_type,
        "sheet_name": parsed["sheet_name"],
        "rows": len(df),
        "source_url": parsed.get("source_url"),
        "last_refreshed": parsed.get("last_refreshed"),
    }
    if dashboard_type == "revenue_bridge":
        weeks = parsed["weeks"]
        latest_week_col = f"W{weeks[-1]}"
        months, quarters = build_periods(weeks)
        base.update({
            "weeks": weeks,
            "months": months,
            "quarters": quarters,
            "teams": tab_counts(df, "team", latest_week_col),
            "segments": tab_counts(df, "segment", latest_week_col),
            "rhs": tab_counts(df, "rh", latest_week_col),
            "cities": tab_counts(df, "city", latest_week_col),
            "regions": tab_counts(df, "region", latest_week_col),
            "products": tab_counts(df, "product", latest_week_col),
            "has_received": parsed["has_received"],
            "has_spillover": parsed["has_spillover"],
        })
    else:
        base.update({
            "columns": parsed["columns"],
            "overview_widgets": parsed["overview_widgets"],
        })
    return base


async def broadcast_shared_update(summary: dict):
    for q in list(SHARED_SUBSCRIBERS):
        try:
            q.put_nowait(summary)
        except asyncio.QueueFull:
            pass  # a slow client will just catch up on its next event


async def _refresh_shared_session() -> dict:
    """Core refresh logic shared by the manual endpoint and the background
    auto-refresh loop. Runs the blocking network/parse work in a thread so
    it never stalls the event loop for other requests."""
    session = RB_SESSIONS.get(SHARED_SESSION_ID)
    if not session or not session.get("source_url"):
        raise ValueError("No shared Google Sheet is linked yet.")

    content = await asyncio.to_thread(fetch_gsheet_as_xlsx, session["source_url"])
    parsed = await asyncio.to_thread(load_workbook_smart, content)
    parsed["source_url"] = session["source_url"]
    parsed["last_refreshed"] = time.time()
    parsed["query_cache"] = {}
    RB_SESSIONS[SHARED_SESSION_ID] = parsed
    return parsed


async def _auto_refresh_loop():
    """Keeps the shared dashboard fresh in the background so viewers don't
    have to trigger a refresh themselves — pushed out over SSE as soon as
    new data lands."""
    while True:
        await asyncio.sleep(SHARED_AUTO_REFRESH_SECONDS)
        try:
            parsed = await _refresh_shared_session()
        except Exception:
            continue  # keep the last good data; try again next tick
        summary = _shared_summary(parsed, SHARED_SESSION_ID)
        await broadcast_shared_update(summary)


@app.on_event("startup")
async def _on_startup():
    asyncio.create_task(_auto_refresh_loop())


@app.get("/api/rb/shared/status")
async def rb_shared_status():
    """Called on page load. Returns the team-wide dashboard if one has been set up, else null."""
    session = RB_SESSIONS.get(SHARED_SESSION_ID)
    if not session:
        return {"active": False}
    return {"active": True, **_shared_summary(session, SHARED_SESSION_ID)}


@app.get("/api/rb/shared/stream")
async def rb_shared_stream():
    """Server-Sent Events stream: pushes a fresh summary to this client
    whenever the shared dashboard refreshes (background loop or manual),
    replacing the old fixed 60s client-side poll."""
    async def event_gen():
        q = asyncio.Queue(maxsize=5)
        SHARED_SUBSCRIBERS.append(q)
        try:
            session = RB_SESSIONS.get(SHARED_SESSION_ID)
            if session:
                yield f"data: {json.dumps(_shared_summary(session, SHARED_SESSION_ID))}\n\n"
            while True:
                data = await q.get()
                yield f"data: {json.dumps(data)}\n\n"
        finally:
            if q in SHARED_SUBSCRIBERS:
                SHARED_SUBSCRIBERS.remove(q)

    return StreamingResponse(event_gen(), media_type="text/event-stream")


@app.post("/api/rb/shared/set")
async def rb_shared_set(payload: dict = Body(...)):
    """Points the team-wide dashboard at a Google Sheet link. Anyone on the team can call this."""
    url = payload.get("sheet_url", "").strip()
    if not url:
        raise HTTPException(status_code=400, detail="Paste a Google Sheets link first.")

    try:
        content = await asyncio.to_thread(fetch_gsheet_as_xlsx, url)
        parsed = await asyncio.to_thread(load_workbook_smart, content)
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))
    except Exception as e:
        raise HTTPException(status_code=502, detail=f"Couldn't load this sheet: {e}")

    parsed["source_url"] = url
    parsed["last_refreshed"] = time.time()
    parsed["query_cache"] = {}
    RB_SESSIONS[SHARED_SESSION_ID] = parsed
    summary = _shared_summary(parsed, SHARED_SESSION_ID)
    await broadcast_shared_update(summary)
    return summary


@app.post("/api/rb/shared/refresh")
async def rb_shared_refresh():
    """Re-pulls the latest data from the currently linked Google Sheet, in place,
    and pushes it to every connected client over SSE."""
    try:
        parsed = await _refresh_shared_session()
    except ValueError as e:
        raise HTTPException(status_code=404, detail=str(e))
    except Exception as e:
        raise HTTPException(status_code=502, detail=f"Couldn't refresh: {e}")

    summary = _shared_summary(parsed, SHARED_SESSION_ID)
    await broadcast_shared_update(summary)
    return summary


@app.post("/api/rb/upload")
async def rb_upload(file: UploadFile = File(...)):
    content = await file.read()
    try:
        parsed = load_dataset_smart(content, file.filename or "")
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))

    parsed["query_cache"] = {}
    rb_session_id = str(uuid.uuid4())
    RB_SESSIONS[rb_session_id] = parsed
    return _shared_summary(parsed, rb_session_id)


def _cached_bridge(session: dict, prev_week: int, curr_week: int, filters: dict, metric: str):
    cache = session.setdefault("query_cache", {})
    key = (int(prev_week), int(curr_week), metric, tuple(sorted(filters.items())))
    if key in cache:
        return cache[key]
    result = compute_metric_bridge(session["df"], int(prev_week), int(curr_week), filters, metric)
    cache[key] = result
    return result


@app.post("/api/rb/query")
async def rb_query(payload: dict = Body(...)):
    rb_session_id = payload.get("rb_session_id")
    prev_week = payload.get("prev_week")
    curr_week = payload.get("curr_week")
    metric = payload.get("metric", "revenue")
    filters = {
        "team": payload.get("team", "All Merchants"),
        "segment": payload.get("segment", "All"),
        "city": payload.get("city", "All"),
        "region": payload.get("region", "All"),
        "product": payload.get("product", "All"),
        "rh": payload.get("rh", "All Merchants"),
    }

    if not rb_session_id or rb_session_id not in RB_SESSIONS:
        raise HTTPException(status_code=404, detail="Session not found. Upload the sheet first.")
    session = RB_SESSIONS[rb_session_id]
    if session.get("dashboard_type") != "revenue_bridge":
        raise HTTPException(status_code=400, detail="This sheet doesn't match the weekly revenue bridge format. Use /api/generic/ask instead.")
    if prev_week is None or curr_week is None:
        raise HTTPException(status_code=400, detail="prev_week and curr_week are required.")

    try:
        result = _cached_bridge(session, prev_week, curr_week, filters, metric)
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))

    return result


@app.post("/api/rb/trend")
async def rb_trend(payload: dict = Body(...)):
    rb_session_id = payload.get("rb_session_id")
    metric = payload.get("metric", "revenue")
    filters = {
        "team": payload.get("team", "All Merchants"),
        "segment": payload.get("segment", "All"),
        "city": payload.get("city", "All"),
        "region": payload.get("region", "All"),
        "product": payload.get("product", "All"),
        "rh": payload.get("rh", "All Merchants"),
    }

    if not rb_session_id or rb_session_id not in RB_SESSIONS:
        raise HTTPException(status_code=404, detail="Session not found. Upload the sheet first.")
    session = RB_SESSIONS[rb_session_id]
    if session.get("dashboard_type") != "revenue_bridge":
        raise HTTPException(status_code=400, detail="Trend view is only available for the weekly revenue bridge format.")

    points = compute_trend(session["df"], session["weeks"], filters, metric)
    return {"metric": metric, "points": points}


@app.post("/api/rb/export")
async def rb_export(payload: dict = Body(...)):
    rb_session_id = payload.get("rb_session_id")
    prev_week = payload.get("prev_week")
    curr_week = payload.get("curr_week")
    metric = payload.get("metric", "revenue")
    filters = {
        "team": payload.get("team", "All Merchants"),
        "segment": payload.get("segment", "All"),
        "city": payload.get("city", "All"),
        "region": payload.get("region", "All"),
        "product": payload.get("product", "All"),
        "rh": payload.get("rh", "All Merchants"),
    }

    if not rb_session_id or rb_session_id not in RB_SESSIONS:
        raise HTTPException(status_code=404, detail="Session not found. Upload the sheet first.")
    session = RB_SESSIONS[rb_session_id]
    if session.get("dashboard_type") != "revenue_bridge":
        raise HTTPException(status_code=400, detail="Export is only available for the weekly revenue bridge format.")
    if prev_week is None or curr_week is None:
        raise HTTPException(status_code=400, detail="prev_week and curr_week are required.")

    try:
        result = _cached_bridge(session, prev_week, curr_week, filters, metric)
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))

    out_df = pd.DataFrame(result["table_rows"])
    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        out_df.to_excel(writer, index=False, sheet_name="export")
    buf.seek(0)
    filename = f"dotdash_export_W{prev_week}_W{curr_week}_{metric}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


def fuzzy_match_category(value, options: list, all_label: str, threshold: int = 70) -> str:
    """Corrects an LLM-parsed category name against the sheet's actual values.
    LLMs often paraphrase or slightly mis-type category names (e.g. 'farming'
    vs 'Farming Team'); this maps back to the closest real value instead of
    silently falling through to 'All' on an exact-match miss."""
    if value is None:
        return all_label
    value_str = str(value).strip()
    if not value_str or value_str.lower() in ("all", all_label.lower()):
        return all_label
    if not options:
        return all_label
    match = process.extractOne(value_str, options, scorer=fuzz.WRatio)
    if match and match[1] >= threshold:
        return match[0]
    return all_label


RB_NL_SYSTEM_TEMPLATE = """You control a merchant revenue bridge dashboard. Output ONLY a JSON object (no markdown fences) with:
- "metric": one of "revenue", "merchants", "stores" — what the user wants to see (defaults to "revenue" if unclear)
- "team": one of [{teams}] or "All Merchants" if unspecified
- "segment": one of [{segments}] or "All" if unspecified
- "rh": one of [{rhs}] or "All Merchants" if unspecified
- "city": one of [{cities}] or "All" if unspecified
- "region": one of [{regions}] or "All" if unspecified
- "product": one of [{products}] or "All" if unspecified
- "prev_week": integer, one of [{weeks}]
- "curr_week": integer, one of [{weeks}]

Rules:
- If the user doesn't mention a week/month/quarter, default to prev_week={default_prev}, curr_week={default_curr}.
- "W0 to W2" or "W0->W2" means prev_week=0, curr_week=2.
- Month/Quarter shortcuts (map to week numbers using this table): {periods}
- Match names case-insensitively to the closest option in the lists above; if nothing matches, use the defaults.
"""


@app.post("/api/rb/nl_query")
async def rb_nl_query(payload: dict = Body(...)):
    rb_session_id = payload.get("rb_session_id")
    prompt = payload.get("prompt", "").strip()
    provider = payload.get("provider", "").strip().lower()
    api_key = payload.get("api_key", "").strip()
    model = payload.get("model", "").strip()

    if not rb_session_id or rb_session_id not in RB_SESSIONS:
        raise HTTPException(status_code=404, detail="Session not found. Upload the sheet first.")
    if RB_SESSIONS[rb_session_id].get("dashboard_type") != "revenue_bridge":
        raise HTTPException(status_code=400, detail="This sheet doesn't match the weekly revenue bridge format. Use /api/generic/ask instead.")
    if not prompt:
        raise HTTPException(status_code=400, detail="Prompt is empty.")
    if provider not in DEFAULT_MODELS:
        raise HTTPException(status_code=400, detail="Pick a provider: anthropic, openai, or gemini.")
    if not api_key:
        raise HTTPException(status_code=400, detail="Add your API key in the settings panel first.")

    session = RB_SESSIONS[rb_session_id]
    df = session["df"]
    weeks = session["weeks"]
    months, quarters = build_periods(weeks)

    def uniq(field):
        return sorted(df[field].unique().tolist()) if field in df.columns else []

    teams, segments, cities, regions, products = uniq("team"), uniq("segment"), uniq("city"), uniq("region"), uniq("product")
    rhs = uniq("rh")
    default_prev = weeks[-2] if len(weeks) >= 2 else weeks[0]
    default_curr = weeks[-1]
    periods_desc = "; ".join(
        [f"{m['label']}=W{m['prev_week']}-W{m['curr_week']}" for m in months] +
        [f"{q['label']}=W{q['prev_week']}-W{q['curr_week']}" for q in quarters]
    ) or "none defined"

    cache_key = (rb_session_id, prompt.lower(), provider, model)
    intent = nl_cache_get(cache_key)

    if intent is None:
        system_prompt = RB_NL_SYSTEM_TEMPLATE.format(
            teams=", ".join(teams) or "none available",
            segments=", ".join(segments) or "none available",
            rhs=", ".join(rhs) or "none available",
            cities=", ".join(cities) or "none available",
            regions=", ".join(regions) or "none available",
            products=", ".join(products) or "none available",
            weeks=", ".join(str(w) for w in weeks),
            default_prev=default_prev,
            default_curr=default_curr,
            periods=periods_desc,
        )
        try:
            raw = call_llm(provider, api_key, model, system_prompt, prompt).strip()
            raw = raw.removeprefix("```json").removeprefix("```").removesuffix("```").strip()
            intent = json.loads(raw)
        except json.JSONDecodeError as e:
            raise HTTPException(status_code=502, detail=f"Model didn't return valid JSON: {e}")
        except Exception as e:
            raise HTTPException(status_code=502, detail=f"{provider} call failed — check your API key. ({e})")
        nl_cache_set(cache_key, intent)

    metric = intent.get("metric") or "revenue"
    if metric not in ("revenue", "merchants", "stores"):
        metric = "revenue"

    # Fuzzy-correct category names instead of requiring an exact match to
    # what the sheet actually contains — LLMs paraphrase category names often.
    filters = {
        "team": fuzzy_match_category(intent.get("team"), teams, "All Merchants"),
        "segment": fuzzy_match_category(intent.get("segment"), segments, "All"),
        "rh": fuzzy_match_category(intent.get("rh"), rhs, "All Merchants"),
        "city": fuzzy_match_category(intent.get("city"), cities, "All"),
        "region": fuzzy_match_category(intent.get("region"), regions, "All"),
        "product": fuzzy_match_category(intent.get("product"), products, "All"),
    }
    prev_week = intent.get("prev_week", default_prev)
    curr_week = intent.get("curr_week", default_curr)

    try:
        result = _cached_bridge(session, prev_week, curr_week, filters, metric)
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))

    result["parsed_intent"] = {**filters, "metric": metric, "prev_week": int(prev_week), "curr_week": int(curr_week)}
    return result


@app.post("/api/generic/ask")
async def generic_ask(payload: dict = Body(...)):
    rb_session_id = payload.get("rb_session_id")
    prompt = payload.get("prompt", "").strip()
    provider = payload.get("provider", "").strip().lower()
    api_key = payload.get("api_key", "").strip()
    model = payload.get("model", "").strip()

    if not rb_session_id or rb_session_id not in RB_SESSIONS:
        raise HTTPException(status_code=404, detail="Session not found. Load a sheet first.")
    session = RB_SESSIONS[rb_session_id]
    if session.get("dashboard_type") != "generic":
        raise HTTPException(status_code=400, detail="This is a revenue bridge sheet — use the week/team controls or the regular prompt box instead.")
    if not prompt:
        raise HTTPException(status_code=400, detail="Question is empty.")
    if provider not in DEFAULT_MODELS:
        raise HTTPException(status_code=400, detail="Pick a provider: anthropic, openai, or gemini.")
    if not api_key:
        raise HTTPException(status_code=400, detail="Add your API key in the settings panel first.")

    df = session["df"]
    schema = get_schema_summary(df)
    user_msg = f"DataFrame schema ({len(df)} rows):\n{schema}\n\nQuestion: {prompt}"

    cache_key = ("generic", rb_session_id, prompt.strip().lower(), provider, model)
    plan = nl_cache_get(cache_key)

    if plan is None:
        try:
            raw = call_llm(provider, api_key, model, GENERIC_SYSTEM_PROMPT, user_msg).strip()
            raw = raw.removeprefix("```json").removeprefix("```").removesuffix("```").strip()
            plan = json.loads(raw)
        except json.JSONDecodeError as e:
            raise HTTPException(status_code=502, detail=f"Model didn't return valid JSON: {e}")
        except Exception as e:
            raise HTTPException(status_code=502, detail=f"{provider} call failed — check your API key. ({e})")
        nl_cache_set(cache_key, plan)

    code = plan.get("pandas_code", "")
    try:
        result = safe_exec_generic(code, df)
        result_json = generic_result_to_json(result)
    except Exception:
        return {
            "explanation": "I generated an analysis but it failed to run. Try rephrasing the question.",
            "chart_type": "none", "data": None, "code": code,
        }

    return {
        "explanation": plan.get("explanation", ""),
        "chart_type": plan.get("chart_type", "none"),
        "chart_x": plan.get("chart_x"),
        "chart_y": plan.get("chart_y"),
        "data": result_json,
        "code": code,
    }


# Serve frontend (must be mounted last so /api routes take priority)
app.mount("/", StaticFiles(directory="static", html=True), name="static")
"""
HTML Output Below
"""